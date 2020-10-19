import cv2
import os
import  numpy as np
from PIL import Image, ImageDraw, ImageFont

from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

dir_lvshi="/lvshitou/"
dir_he_1 = "/he_1/"
dir_Final = "/Final/"
beijing_tu = "/beijing.png"

def lvshi(touxiang,name,Law,content):
    if os.access(dir_lvshi+name+".png", os.F_OK):
        img=cv2.imread(touxiang)
        img_back=cv2.imread(beijing_tu)    #背景

        rows,cols,channels = img.shape#rows，cols最后一定要是前景图片的，后面遍历图片需要用到

        #转换hsv
        hsv=cv2.cvtColor(img,cv2.COLOR_BGR2HSV)
        #获取mask
        lower_blue=np.array([0,0,0])
        upper_blue=np.array([0,0,0])
        mask = cv2.inRange(hsv, lower_blue, upper_blue)
        # cv2.imshow('Mask', mask)

        #腐蚀膨胀
        erode=cv2.erode(mask,None,iterations=1)
        # cv2.imshow('erode',erode)
        dilate=cv2.dilate(erode,None,iterations=1)
        # cv2.imshow('dilate',dilate)

        #遍历替换
        center=[70,56]#在新背景图片中的位置 # 律师头像位置
        # center=[154,142] #认证图片位置
        for i in range(rows):
            for j in range(cols):
                if dilate[i,j]==0:#0代表黑色的点
                    img_back[center[0]+i,center[1]+j]=img[i,j]#此处替换颜色，为BGR通道
        cv2.imwrite(dir_he_1+name+".png", img_back)

        tianjiawenzi(dir_he_1+name+".png",name,Law,content)
    else:
        os.system("echo "+name+">> /error.txt")

def tianjiawenzi(dirfile,name,Law,content):
    img = Image.open(dirfile)
    draw = ImageDraw.Draw(img)

    ttfront = ImageFont.truetype('/简体/PingFang Heavy.ttf',34)#字体大小
    ttfront2 = ImageFont.truetype('/简体/PingFang Light.ttf',28)#字体大小
    ttfront3 = ImageFont.truetype('/简体/PingFang Light.ttf',26)#字体大小
    ttfront4 = ImageFont.truetype('/简体/PingFang Heavy.ttf',30)#字体大小
    if (len(name) == 2 and len(content) <= 15):
        draw.text((208, 66),(name),fill=(255,255,255), font=ttfront)#文字位置，内容，字体
        draw.text((280, 73),("律师"),fill=(255,255,255), font=ttfront2)
        draw.text((208, 108),(Law),fill=(255,255,255), font=ttfront3)
        draw.text((208, 150),(content),fill=(255,255,255), font=ttfront4)
        img.save(dir_Final+name+"_"+content+".jpg","JPEG")
        img.show()
    if (len(name) == 3 and len(content) <= 15):
        draw.text((208, 66),(name),fill=(255,255,255), font=ttfront)#文字位置，内容，字体
        draw.text((315, 73),("律师"),fill=(255,255,255), font=ttfront2)
        draw.text((208, 108),(Law),fill=(255,255,255), font=ttfront3)
        draw.text((208, 150),(content),fill=(255,255,255), font=ttfront4)
        img.save(dir_Final+name+"_"+content+".jpg","JPEG")
        img.show()
    if (len(name) == 3 and len(content) > 15):
        draw.text((208, 66),(name),fill=(255,255,255), font=ttfront)#文字位置，内容，字体
        draw.text((315, 73),("律师"),fill=(255,255,255), font=ttfront2)
        draw.text((208, 108),(Law),fill=(255,255,255), font=ttfront3)
        f_name = (content[:15]+"\n"+content[15:])
        draw.text((208, 150),(f_name),fill=(255,255,255), font=ttfront4)
        img.save(dir_Final+name+"_"+content+".jpg","JPEG")
        img.show()
    if (len(name) == 2 and len(content) > 15):
        draw.text((208, 66),(name),fill=(255,255,255), font=ttfront)#文字位置，内容，字体
        draw.text((280, 73),("律师"),fill=(255,255,255), font=ttfront2)
        draw.text((208, 108),(Law),fill=(255,255,255), font=ttfront3)
        f_name = (content[:15]+"\n"+content[15:])
        draw.text((208, 150),(f_name),fill=(255,255,255), font=ttfront4)
        img.save(dir_Final+name+"_"+content+".jpg","JPEG")
        img.show()

if __name__ == '__main__':
    wb=load_workbook('批量处理.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    wb = Workbook()
    sheet = wb.active
    for a in range(ws.max_row):
        # print((dir_lvshi+ws.cell(a,8).value+".png",ws.cell(a,8).value,ws.cell(a,11).value,ws.cell(a,4).value))
        lvshi(dir_lvshi+ws.cell(a+1,8).value+".png",ws.cell(a+1,8).value,ws.cell(a+1,11).value,ws.cell(a+1,4).value)
